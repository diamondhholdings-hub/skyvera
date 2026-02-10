#!/usr/bin/env python3
"""
Enhanced Excel parser that outputs comprehensive JSON via stdout.
Extracts customer data and financial summaries from Skyvera budget file.

Usage:
  python3 parse_excel_to_json.py --type customers|financials|all

Output: JSON to stdout (progress messages to stderr)
"""

import sys
import json
import argparse
import re
from pathlib import Path
from openpyxl import load_workbook

# Excel file path relative to project root
EXCEL_FILE = "2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx"

def log(message):
    """Log progress to stderr (so it doesn't contaminate JSON output)"""
    print(message, file=sys.stderr)

def extract_rr_customers(wb, company_filter):
    """Extract RR customers from 'RR Input' sheet for a specific company."""
    try:
        ws = wb['RR Input']
    except KeyError:
        log(f"Warning: 'RR Input' sheet not found")
        return {}

    customers = {}

    for row in ws.iter_rows(min_row=11, values_only=True):
        if not row or not row[0]:
            continue

        company = row[0]
        customer_name = row[1]
        sub_id = row[3] if len(row) > 3 else None
        arr = row[6] if len(row) > 6 else 0
        renewal_qtr = row[8] if len(row) > 8 else None
        will_renew = row[9] if len(row) > 9 else None
        projected_arr = row[11] if len(row) > 11 else None

        if company != company_filter or not customer_name:
            continue

        if customer_name not in customers:
            customers[customer_name] = {
                'customer_name': customer_name,
                'rr': 0,
                'nrr': 0,
                'subscriptions': []
            }

        customers[customer_name]['rr'] += arr if arr else 0
        customers[customer_name]['subscriptions'].append({
            'sub_id': sub_id,
            'arr': arr if arr else 0,
            'renewal_qtr': renewal_qtr,
            'will_renew': will_renew,
            'projected_arr': projected_arr
        })

    return customers

def extract_nrr_customers(wb, class_filter):
    """Extract NRR customers from 'NRR Input' sheet for a specific class."""
    try:
        ws = wb['NRR Input']
    except KeyError:
        log(f"Warning: 'NRR Input' sheet not found")
        return {}

    customers = {}

    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or not row[1]:
            continue

        customer_name = row[1]
        class_col = row[2] if len(row) > 2 else None
        q1_26 = row[8] if len(row) > 8 else 0
        q2_26 = row[9] if len(row) > 9 else 0
        q3_26 = row[10] if len(row) > 10 else 0
        q4_26 = row[11] if len(row) > 11 else 0

        if not class_col or class_filter not in class_col:
            continue

        # Extract customer name from <name> format
        match = re.search(r'<(.+?)>', customer_name) if customer_name else None
        if match:
            customer_name = match.group(1).strip()

        if not customer_name:
            continue

        fy26_nrr = (q1_26 or 0) + (q2_26 or 0) + (q3_26 or 0) + (q4_26 or 0)

        if customer_name not in customers:
            customers[customer_name] = 0
        customers[customer_name] += fy26_nrr

    return customers

def extract_bu_customers(wb, bu_name, company_filter, class_filter):
    """Extract all customers for a BU, combining RR and NRR data."""
    log(f"Extracting {bu_name}...")

    rr_customers = extract_rr_customers(wb, company_filter)
    nrr_data = extract_nrr_customers(wb, class_filter)

    # Merge RR and NRR data
    all_customer_names = set(rr_customers.keys()) | set(nrr_data.keys())

    customers = []
    for name in all_customer_names:
        rr = rr_customers.get(name, {}).get('rr', 0)
        nrr = nrr_data.get(name, 0)
        total = rr + nrr

        customer = {
            'customer_name': name,
            'rr': rr,
            'nrr': nrr,
            'total': total,
            'subscriptions': rr_customers.get(name, {}).get('subscriptions', [])
        }
        customers.append(customer)

    # Sort by total revenue descending
    customers.sort(key=lambda x: x['total'], reverse=True)

    # Add rank and percentage
    total_revenue = sum(c['total'] for c in customers)
    for i, customer in enumerate(customers, 1):
        customer['rank'] = i
        customer['pct_of_total'] = (customer['total'] / total_revenue * 100) if total_revenue > 0 else 0

    log(f"  ✓ {len(customers)} customers, ${total_revenue:,.0f} total revenue")

    return customers

def extract_all_customers(wb):
    """Extract customers from all BUs."""
    bu_configs = [
        ('Cloudsense', 'Cloudsense', 'Cloudsense'),
        ('Kandy', 'Kandy', 'Kandy'),
        ('STL', 'STL', 'Stl'),
        ('NewNet', 'NewNet', 'Newnet')
    ]

    customers_by_bu = {}

    for bu_name, company_filter, class_filter in bu_configs:
        customers = extract_bu_customers(wb, bu_name, company_filter, class_filter)
        customers_by_bu[bu_name] = customers

    return customers_by_bu

def _safe_num(value, default=0):
    """Safely convert a cell value to a float, returning default if None or non-numeric."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _read_bu_pnl(wb, sheet_name, bu_name):
    """Read actual P&L values from a BU-specific P&L sheet.

    All three BU P&L sheets (Cloudsense, Kandy, STL) share an identical
    layout.  The Q1'26 BU Plan values are in column C (index 3).

    Row mapping (column B = label, column C = Q1'26 BU Plan value):
        Row  5: Recurring Revenue
        Row  6: Non Recurring Revenue
        Row  7: Total Revenue
        Row  8: HC COGS
        Row  9: NHC COGS
        Row 10: CF COGS
        Row 11: Total COGS
        Row 12: Gross Profit
        Row 13: Gross Margin  (decimal, e.g. 0.74)
        Row 14: HC Expenses
        Row 15: NHC Expenses
        Row 16: Total Revenue Write Off
        Row 17: CF Expenses
        Row 18: Core Allocation
        Row 19: Total Expenses
        Row 20: Net Profit
        Row 21: Net Margin  (decimal, e.g. 0.59)
        Row 22: Margin Target  (decimal)
        Row 23: Delta to Margin
        Row 24: EBITDA
    """
    try:
        ws = wb[sheet_name]
    except KeyError:
        log(f"  Warning: Sheet '{sheet_name}' not found, skipping {bu_name}")
        return None

    COL = 3  # Column C = Q1'26 BU Plan

    def cell(row):
        return _safe_num(ws.cell(row=row, column=COL).value)

    rr             = cell(5)
    nrr            = cell(6)
    total_revenue  = cell(7)
    hc_cogs        = cell(8)
    nhc_cogs       = cell(9)
    cf_cogs        = cell(10)
    total_cogs     = cell(11)
    gross_profit   = cell(12)
    gross_margin   = cell(13)
    hc_expenses    = cell(14)
    nhc_expenses   = cell(15)
    rev_write_off  = cell(16)
    cf_expenses    = cell(17)
    core_alloc     = cell(18)
    total_expenses = cell(19)
    net_profit     = cell(20)
    net_margin_dec = cell(21)
    margin_target  = cell(22)
    delta_to_margin = cell(23)
    ebitda         = cell(24)

    return {
        'bu': bu_name,
        'totalRR': rr,
        'totalNRR': nrr,
        'totalRevenue': total_revenue,
        'hcCogs': hc_cogs,
        'nhcCogs': nhc_cogs,
        'cfCogs': cf_cogs,
        'totalCogs': total_cogs,
        'grossProfit': gross_profit,
        'grossMargin': gross_margin * 100,        # convert to percentage
        'hcExpenses': hc_expenses,
        'nhcExpenses': nhc_expenses,
        'revenueWriteOff': rev_write_off,
        'cfExpenses': cf_expenses,
        'coreAllocation': core_alloc,
        'totalExpenses': total_expenses,
        'netProfit': net_profit,
        'netMargin': net_margin_dec * 100,         # convert to percentage
        'marginTarget': margin_target * 100,       # convert to percentage
        'deltaToMargin': delta_to_margin,
        'ebitda': ebitda,
    }


def extract_financials(wb):
    """Extract actual financial data from BU-specific P&L sheets.

    Reads directly from the 'P&Ls - <BU>' sheets instead of estimating
    costs with hardcoded ratios.  Also enriches each BU record with
    customer count from the RR/NRR input sheets.
    """
    log("Extracting financial summaries from P&L sheets...")

    # Map of BU name -> (P&L sheet name, RR company filter, NRR class filter)
    bu_configs = {
        'Cloudsense': ("P&Ls - Cloudsense", 'Cloudsense', 'Cloudsense'),
        'Kandy':      ("P&Ls - Kandy",      'Kandy',      'Kandy'),
        'STL':        ("P&Ls - STL",         'STL',        'Stl'),
    }

    # Get customer counts per BU (reuse existing extraction logic)
    customers_by_bu = extract_all_customers(wb)

    financials_by_bu = {}

    for bu_name, (sheet_name, _, _) in bu_configs.items():
        pnl = _read_bu_pnl(wb, sheet_name, bu_name)
        if pnl is None:
            continue

        # Attach customer count
        pnl['customerCount'] = len(customers_by_bu.get(bu_name, []))

        financials_by_bu[bu_name] = pnl
        log(f"  ✓ {bu_name}: ${pnl['totalRevenue']:,.0f} revenue, "
            f"{pnl['netMargin']:.1f}% net margin (from {sheet_name})")

    # Also read the consolidated Skyvera totals from the 'P&Ls' sheet
    consolidated = _read_bu_pnl(wb, "P&Ls", "Skyvera")
    if consolidated:
        total_customers = sum(
            len(customers_by_bu.get(bu, []))
            for bu in bu_configs
        )
        consolidated['customerCount'] = total_customers
        financials_by_bu['Skyvera'] = consolidated
        log(f"  ✓ Skyvera (consolidated): ${consolidated['totalRevenue']:,.0f} revenue, "
            f"{consolidated['netMargin']:.1f}% net margin")

    return financials_by_bu

def main():
    parser = argparse.ArgumentParser(description='Parse Skyvera Excel budget file to JSON')
    parser.add_argument('--type', choices=['customers', 'financials', 'all'], required=True,
                        help='Type of data to extract')
    args = parser.parse_args()

    # Check if Excel file exists
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        log(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    log(f"Loading workbook: {EXCEL_FILE}")
    log("This may take 10-15 seconds...")

    try:
        # Load workbook with data_only=True to get calculated values
        wb = load_workbook(EXCEL_FILE, data_only=True)
        log("✓ Workbook loaded")
    except Exception as e:
        log(f"ERROR: Failed to load workbook: {e}")
        sys.exit(1)

    # Extract data based on type
    result = {}

    if args.type in ['customers', 'all']:
        customers = extract_all_customers(wb)
        if args.type == 'customers':
            result = {'customers': customers}
        else:
            result['customers'] = customers

    if args.type in ['financials', 'all']:
        financials = extract_financials(wb)
        if args.type == 'financials':
            result = {'financials': financials}
        else:
            result['financials'] = financials

    # Output JSON to stdout
    log("\n✓ Extraction complete, outputting JSON...")
    print(json.dumps(result, indent=2))

    # Summary to stderr
    if 'customers' in result:
        total_customers = sum(len(customers) for customers in result['customers'].values())
        log(f"\nTotal customers extracted: {total_customers}")
    if 'financials' in result:
        log(f"Financial summaries: {len(result['financials'])} BUs")

if __name__ == '__main__':
    main()
