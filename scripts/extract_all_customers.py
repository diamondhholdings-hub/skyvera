#!/usr/bin/env python3
"""Extract ALL customers (100%) for analytics dashboard."""

import json
from openpyxl import load_workbook
import re

def extract_all_rr(company_filter):
    """Extract all RR for a specific company."""
    wb = load_workbook("2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx", data_only=True)
    ws = wb['RR Input']
    customers = {}

    for row in ws.iter_rows(min_row=11, values_only=True):
        company = row[0]
        customer_name = row[1]
        sub_id = row[3]
        arr = row[6]
        renewal_qtr = row[8]
        will_renew = row[9]
        projected_arr = row[11]

        if company != company_filter or not customer_name:
            continue

        if customer_name not in customers:
            customers[customer_name] = {
                'customer_name': customer_name,
                'rr': 0,
                'nrr': 0,
                'subscriptions': []
            }

        customers[customer_name]['rr'] += arr if arr else 0
        customers[customer_name]['subscriptions'].append({
            'sub_id': sub_id,
            'arr': arr,
            'renewal_qtr': renewal_qtr,
            'will_renew': will_renew,
            'projected_arr': projected_arr
        })

    return customers

def extract_all_nrr(class_filter):
    """Extract all NRR for a specific class."""
    wb = load_workbook("2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx", data_only=True)
    ws = wb['NRR Input']
    customers = {}

    for row in ws.iter_rows(min_row=6, values_only=True):
        customer_name = row[1]
        class_col = row[2]
        q1_26 = row[8] if len(row) > 8 else 0
        q2_26 = row[9] if len(row) > 9 else 0
        q3_26 = row[10] if len(row) > 10 else 0
        q4_26 = row[11] if len(row) > 11 else 0

        if not class_col or class_filter not in class_col:
            continue

        match = re.search(r'<(.+?)>', customer_name) if customer_name else None
        if match:
            customer_name = match.group(1).strip()

        if not customer_name:
            continue

        fy26_nrr = (q1_26 or 0) + (q2_26 or 0) + (q3_26 or 0) + (q4_26 or 0)

        if customer_name not in customers:
            customers[customer_name] = 0
        customers[customer_name] += fy26_nrr

    return customers

def extract_bu(bu_name, company_filter, class_filter):
    """Extract all customers for a BU."""
    rr_customers = extract_all_rr(company_filter)
    nrr_data = extract_all_nrr(class_filter)

    all_customer_names = set(rr_customers.keys()) | set(nrr_data.keys())

    customers = []
    for name in all_customer_names:
        rr = rr_customers.get(name, {}).get('rr', 0)
        nrr = nrr_data.get(name, 0)
        total = rr + nrr

        customer = {
            'customer_name': name,
            'rr': rr,
            'nrr': nrr,
            'total': total,
            'subscriptions': rr_customers.get(name, {}).get('subscriptions', [])
        }
        customers.append(customer)

    customers.sort(key=lambda x: x['total'], reverse=True)

    # Add rank and percentage
    total_revenue = sum(c['total'] for c in customers)
    for i, customer in enumerate(customers, 1):
        customer['rank'] = i
        customer['pct_of_total'] = (customer['total'] / total_revenue * 100) if total_revenue > 0 else 0

    return {
        'bu_name': bu_name,
        'total_revenue': total_revenue,
        'customer_count': len(customers),
        'customers': customers
    }

def main():
    print("="*80)
    print("EXTRACTING ALL CUSTOMERS (100%) FOR ANALYTICS")
    print("="*80)

    bu_configs = [
        ('CloudSense', 'Cloudsense', 'Cloudsense'),
        ('Kandy', 'Kandy', 'Kandy'),
        ('STL', 'STL', 'Stl'),
        ('NewNet', 'NewNet', 'Newnet')
    ]

    for bu_name, company_filter, class_filter in bu_configs:
        print(f"\nExtracting {bu_name}...")
        data = extract_bu(bu_name, company_filter, class_filter)
        
        filename = f'data/customers_{bu_name.lower()}_all.json'
        with open(filename, 'w') as f:
            json.dump(data, f, indent=2)
        
        print(f"  ✅ {data['customer_count']} customers, ${data['total_revenue']:,.0f} total revenue")
        print(f"  📁 Saved to: {filename}")

    print("\n" + "="*80)
    print("✅ All customer data extracted for analytics")
    print("="*80)

if __name__ == '__main__':
    main()
