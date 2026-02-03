"""Extract top 80% NewNet customers by total revenue (RR+NRR)."""
import json
from openpyxl import load_workbook
from datetime import datetime
import re

def extract_newnet_rr():
    """Extract NewNet RR from RR Input sheet."""
    wb = load_workbook("2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx", data_only=True)
    ws = wb['RR Input']

    customers = {}

    for row in ws.iter_rows(min_row=11, values_only=True):
        company = row[0]  # Column A
        customer_name = row[1]  # Column B
        sub_id = row[3]  # Column D
        arr = row[6]  # Column G
        renewal_qtr = row[8]  # Column I
        will_renew = row[9]  # Column J
        upsell_pct = row[10]  # Column K
        projected_arr = row[11]  # Column L

        if company != 'NewNet' or not customer_name:
            continue

        if customer_name not in customers:
            customers[customer_name] = {
                'customer_name': customer_name,
                'rr': 0,
                'nrr': 0,
                'subscriptions': []
            }

        customers[customer_name]['rr'] += arr if arr else 0
        customers[customer_name]['subscriptions'].append({
            'sub_id': sub_id,
            'arr': arr,
            'renewal_qtr': renewal_qtr,
            'will_renew': will_renew,
            'projected_arr': projected_arr
        })

    return customers

def extract_newnet_nrr():
    """Extract NewNet NRR from NRR Input sheet."""
    wb = load_workbook("2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx", data_only=True)
    ws = wb['NRR Input']

    customers = {}

    for row in ws.iter_rows(min_row=6, values_only=True):
        customer_name = row[1]  # Column B
        class_col = row[2]  # Column C
        q1_26 = row[8] if len(row) > 8 else 0  # Column I
        q2_26 = row[9] if len(row) > 9 else 0  # Column J
        q3_26 = row[10] if len(row) > 10 else 0  # Column K
        q4_26 = row[11] if len(row) > 11 else 0  # Column L

        if not class_col or 'Newnet' not in class_col:
            continue

        # Handle "New Sales Ps - <Customer>" pattern
        match = re.search(r'<(.+?)>', customer_name) if customer_name else None
        if match:
            customer_name = match.group(1).strip()

        if not customer_name:
            continue

        fy26_nrr = (q1_26 or 0) + (q2_26 or 0) + (q3_26 or 0) + (q4_26 or 0)

        if customer_name not in customers:
            customers[customer_name] = 0
        customers[customer_name] += fy26_nrr

    return customers

def main():
    print("="*100)
    print("EXTRACTING NEWNET CUSTOMER DATA")
    print("="*100)

    # Get RR data
    rr_customers = extract_newnet_rr()
    print(f"\nFound {len(rr_customers)} NewNet customers with RR")

    # Get NRR data
    nrr_data = extract_newnet_nrr()
    print(f"Found {len(nrr_data)} NewNet customers with NRR")

    # Merge RR and NRR
    all_customer_names = set(rr_customers.keys()) | set(nrr_data.keys())

    customers = []
    for name in all_customer_names:
        rr = rr_customers.get(name, {}).get('rr', 0)
        nrr = nrr_data.get(name, 0)
        total = rr + nrr

        customer = {
            'customer_name': name,
            'rr': rr,
            'nrr': nrr,
            'total': total,
            'subscriptions': rr_customers.get(name, {}).get('subscriptions', [])
        }
        customers.append(customer)

    # Sort by total revenue descending
    customers.sort(key=lambda x: x['total'], reverse=True)

    # Calculate top 80%
    total_revenue = sum(c['total'] for c in customers)
    cumulative = 0
    top_80_customers = []

    for customer in customers:
        cumulative += customer['total']
        top_80_customers.append(customer)
        if cumulative >= total_revenue * 0.8:
            break

    # Add rank and percentage
    for i, customer in enumerate(top_80_customers, 1):
        customer['rank'] = i
        customer['pct_of_total'] = (customer['total'] / total_revenue * 100) if total_revenue > 0 else 0

    # Save to JSON
    output = {
        'total_revenue': total_revenue,
        'top_80_count': len(top_80_customers),
        'top_80_revenue': sum(c['total'] for c in top_80_customers),
        'customers': top_80_customers
    }

    with open('data/customers_newnet_top80.json', 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\n{'='*100}")
    print(f"✅ Extracted {len(top_80_customers)} NewNet customers (80% of ${total_revenue:,.0f})")
    print(f"📁 Saved to: data/customers_newnet_top80.json")
    print("="*100)

if __name__ == '__main__':
    main()
