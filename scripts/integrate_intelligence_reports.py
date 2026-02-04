#!/usr/bin/env python3
"""
Integrate customer intelligence reports into HTML dashboards.
Parses markdown intelligence reports and generates HTML sections.
"""

import re
import os
from pathlib import Path
import markdown2
from openpyxl import load_workbook
from datetime import datetime

def load_renewal_dates():
    """Extract renewal dates from Excel budget file."""
    renewal_data = {}

    try:
        wb = load_workbook("2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx", data_only=True)
        ws = wb['RR Input']

        # Skip header rows (first 10 rows)
        for row in ws.iter_rows(min_row=11, values_only=True):
            if not row[1]:  # Skip if no customer name
                continue

            customer_name = row[1]  # Column B
            renewal_date = row[7]  # Column H
            renewal_qtr = row[8]    # Column I
            will_renew = row[9]      # Column J

            if customer_name and customer_name not in ['SUM', 'SUBTOTAL']:
                # Normalize customer name for matching
                customer_key = customer_name.replace('/', '-').replace(' ', '_')

                renewal_info = {
                    'renewal_date': renewal_date,
                    'renewal_qtr': renewal_qtr,
                    'will_renew': will_renew
                }

                renewal_data[customer_key] = renewal_info
                # Also store with spaces for alternate matching
                renewal_data[customer_name] = renewal_info

        print(f"✅ Loaded renewal data for {len(set(renewal_data.keys()))} customers")
    except Exception as e:
        print(f"⚠️  Could not load renewal dates: {e}")

    return renewal_data

def parse_intelligence_report(filepath):
    """Parse markdown intelligence report and extract key sections."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Extract customer name from first heading
    customer_match = re.search(r'^#\s*(.+?)(?:\n|$)', content, re.MULTILINE)
    customer_name = customer_match.group(1) if customer_match else "Unknown"
    customer_name = re.sub(r'^Customer Intelligence (Report|Brief):\s*', '', customer_name)

    # Extract all sections (## headers)
    sections = {}
    section_pattern = r'##\s+(.+?)\n(.*?)(?=\n##\s+|\Z)'
    matches = re.findall(section_pattern, content, re.DOTALL)

    for title, section_content in matches:
        sections[title.strip()] = section_content.strip()

    return {
        'customer_name': customer_name,
        'sections': sections,
        'raw_content': content
    }

def extract_health_score(content):
    """Extract strategic health score from anywhere in document."""
    score_match = re.search(r'(?:Strategic Health Score|Health Score):\s*(\d+\.?\d*)\s*/\s*10', content, re.IGNORECASE)
    if score_match:
        return float(score_match.group(1))
    return None

def extract_key_findings(content):
    """Extract key findings from executive summary or key findings section."""
    # Try to find Key Findings Summary or similar sections
    findings_section = re.search(r'###\s*Key Findings.*?\n+(.*?)(?=\n---|\n##|\Z)', content, re.DOTALL)
    if findings_section:
        findings_text = findings_section.group(1).strip()
        # Extract numbered findings with bold titles
        findings = re.findall(r'\d+\.\s+\*\*(.+?)\*\*[:\s]*(.+?)(?=\n\d+\.|\Z)', findings_text, re.DOTALL)
        return [{'title': f[0].strip(':'), 'description': f[1].strip()} for f in findings]
    
    # Alternative: Look for numbered points in EXECUTIVE SUMMARY
    exec_summary = re.search(r'##\s*EXECUTIVE SUMMARY\s*\n(.*?)(?=\n##|\Z)', content, re.DOTALL)
    if exec_summary:
        summary_text = exec_summary.group(1)
        findings = re.findall(r'(?:^|\n)(\d+)\.\s+\*\*(.+?)\*\*[:\s]*(.+?)(?=\n\d+\.|\n##|\Z)', summary_text, re.DOTALL)
        return [{'title': f[1].strip(':'), 'description': f[2].strip()[:200]} for f in findings[:4]]
    
    return []

def markdown_to_html(md_content):
    """Convert markdown to HTML with tables and formatting."""
    html = markdown2.markdown(
        md_content,
        extras=['tables', 'fenced-code-blocks', 'header-ids']
    )

    # Add CSS class to tables for proper styling
    html = html.replace('<table>', '<table class="data-table">')

    # Ensure proper paragraph spacing in list items
    html = html.replace('<li><p>', '<li>').replace('</p></li>', '</li>')

    return html

def generate_intelligence_html(report_data, renewal_info=None):
    """Generate HTML section for intelligence data."""
    sections = report_data['sections']
    raw_content = report_data['raw_content']

    # Extract components
    health_score = extract_health_score(raw_content)
    key_findings = extract_key_findings(raw_content)

    html_parts = []

    # Renewal Information Banner
    if renewal_info:
        renewal_date = renewal_info.get('renewal_date')
        renewal_qtr = renewal_info.get('renewal_qtr')
        will_renew = renewal_info.get('will_renew')

        # Format renewal date
        if isinstance(renewal_date, datetime):
            renewal_date_str = renewal_date.strftime('%B %d, %Y')
        elif renewal_date:
            renewal_date_str = str(renewal_date)
        else:
            renewal_date_str = 'Unknown'

        # Determine status color
        if will_renew == 'Yes':
            status_color = '#4caf50'
            status_icon = '✅'
            status_text = 'RENEWAL EXPECTED'
        elif will_renew == 'No' or will_renew == 'No (SF)':
            status_color = '#e53935'
            status_icon = '❌'
            status_text = 'CHURN RISK'
        else:
            status_color = '#ff9800'
            status_icon = '⚠️'
            status_text = 'RENEWAL UNCERTAIN'

        html_parts.append(f'''
        <div class="card" style="background: linear-gradient(135deg, {status_color}15 0%, {status_color}05 100%); border-left: 4px solid {status_color};">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <div>
                    <div style="font-size: 0.85rem; font-weight: 600; text-transform: uppercase; color: {status_color}; margin-bottom: 0.5rem;">
                        {status_icon} {status_text}
                    </div>
                    <h2 style="margin: 0;">Renewal: {renewal_qtr}</h2>
                    <p style="color: var(--muted); margin-top: 0.5rem;">Contract End Date: {renewal_date_str}</p>
                </div>
                <div style="text-align: right;">
                    <div style="font-size: 3rem; color: {status_color};">{status_icon}</div>
                </div>
            </div>
        </div>
        ''')

    # Health Score Section
    if health_score:
        score_color = '#e65100' if health_score < 5 else '#ff9800' if health_score < 7 else '#4caf50'
        html_parts.append(f'''
        <div class="card" style="background: linear-gradient(135deg, {score_color}15 0%, {score_color}05 100%); border-left: 4px solid {score_color};">
            <h2>📊 Strategic Health Score</h2>
            <div style="display: flex; align-items: center; gap: 2rem; margin-top: 1rem;">
                <div style="font-size: 4rem; font-weight: 700; color: {score_color};">{health_score}/10</div>
                <div style="flex: 1;">
                    <div style="background: #e0e0e0; height: 20px; border-radius: 10px; overflow: hidden;">
                        <div style="background: {score_color}; height: 100%; width: {health_score * 10}%; transition: width 0.3s;"></div>
                    </div>
                    <p style="margin-top: 0.5rem; color: var(--muted); font-size: 0.9rem;">
                        {
                            'CRITICAL - Requires immediate attention' if health_score < 4 else
                            'AT RISK - Monitor closely' if health_score < 6 else
                            'STABLE - Standard engagement' if health_score < 8 else
                            'HEALTHY - Strong relationship'
                        }
                    </p>
                </div>
            </div>
        </div>
        ''')
    
    # Key Findings
    if key_findings:
        findings_html = []
        for idx, finding in enumerate(key_findings[:4], 1):
            icon = '🔴' if any(word in finding['title'].upper() for word in ['CRITICAL', 'RISK', 'THREAT']) else '🟡' if 'CAUTION' in finding['title'].upper() else '🟢'
            findings_html.append(f'''
            <div class="metric-box" style="flex: 1; min-width: 250px;">
                <div style="font-size: 2rem; margin-bottom: 0.5rem;">{icon}</div>
                <div class="label" style="font-weight: 600; color: var(--ink); margin-bottom: 0.5rem;">{finding['title']}</div>
                <div style="font-size: 0.85rem; line-height: 1.5;">{finding['description'][:150]}...</div>
            </div>
            ''')
        
        html_parts.append(f'''
        <div class="card">
            <h2>🎯 Key Strategic Insights</h2>
            <div style="display: flex; flex-wrap: wrap; gap: 1.5rem; margin-top: 1rem;">
                {''.join(findings_html)}
            </div>
        </div>
        ''')
    
    # Look for pain points, competitive, opportunities sections (flexible matching)
    pain_points_keys = [k for k in sections.keys() if 'PAIN POINT' in k.upper() or 'CHALLENGE' in k.upper() or 'STRATEGIC INIT' in k.upper()]
    if pain_points_keys:
        pain_points_md = sections[pain_points_keys[0]]
        pain_points_html = markdown_to_html(pain_points_md)
        html_parts.append(f'''
        <div class="card">
            <h2>⚠️ Pain Points & Strategic Initiatives</h2>
            <div style="margin-top: 1rem;">
                {pain_points_html}
            </div>
        </div>
        ''')
    
    # Competitive Landscape
    competitive_keys = [k for k in sections.keys() if 'COMPETITIVE' in k.upper()]
    if competitive_keys:
        competitive_md = sections[competitive_keys[0]]
        competitive_html = markdown_to_html(competitive_md)
        html_parts.append(f'''
        <div class="card">
            <h2>🏆 Competitive Landscape</h2>
            <div style="margin-top: 1rem;">
                {competitive_html}
            </div>
        </div>
        ''')
    
    # Opportunities
    opportunity_keys = [k for k in sections.keys() if 'OPPORTUNIT' in k.upper() or 'GROWTH' in k.upper() or 'EXPANSION' in k.upper()]
    if opportunity_keys:
        opportunities_md = sections[opportunity_keys[0]]
        opportunities_html = markdown_to_html(opportunities_md)
        html_parts.append(f'''
        <div class="card">
            <h2>🚀 Opportunities & Growth Potential</h2>
            <div style="margin-top: 1rem;">
                {opportunities_html}
            </div>
        </div>
        ''')
    
    # Account Strategy
    strategy_keys = [k for k in sections.keys() if 'STRATEGY' in k.upper() or 'ACTION PLAN' in k.upper() or 'NEXT STEPS' in k.upper()]
    if strategy_keys:
        strategy_md = sections[strategy_keys[0]]
        strategy_html = markdown_to_html(strategy_md)
        html_parts.append(f'''
        <div class="card" style="background: var(--highlight);">
            <h2>📋 Account Strategy & Next Steps</h2>
            <div style="margin-top: 1rem;">
                {strategy_html}
            </div>
        </div>
        ''')
    
    return '\n'.join(html_parts)

def main():
    """Integrate all intelligence reports."""
    reports_dir = Path('data/intelligence/reports')

    if not reports_dir.exists():
        print("❌ No intelligence reports directory found")
        return

    print("=" * 80)
    print("INTEGRATING INTELLIGENCE REPORTS")
    print("=" * 80)
    print()

    # Load renewal dates from Excel
    print("📊 Loading renewal dates from Excel...")
    renewal_data = load_renewal_dates()
    print()

    # Process each report
    report_files = list(reports_dir.glob('*.md'))
    intelligence_data = {}

    for report_file in report_files:
        customer_key = report_file.stem
        print(f"📄 Processing {customer_key}...")

        try:
            report_data = parse_intelligence_report(report_file)
            customer_name = report_data['customer_name']

            # Get renewal info for this customer
            renewal_info = renewal_data.get(customer_key) or renewal_data.get(customer_name)

            html_content = generate_intelligence_html(report_data, renewal_info)
            intelligence_data[customer_key] = html_content
            print(f"   ✅ Successfully parsed ({len(html_content)} chars)")
            if renewal_info:
                print(f"   📅 Renewal: {renewal_info.get('renewal_qtr')} - {renewal_info.get('will_renew')}")
        except Exception as e:
            print(f"   ❌ Error: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    # Save to JSON for dashboard generators to load
    import json
    output_file = 'data/intelligence_html.json'
    with open(output_file, 'w') as f:
        json.dump(intelligence_data, f, indent=2)
    
    print()
    print(f"✅ Processed {len(intelligence_data)} intelligence reports")
    print(f"📁 Saved to: {output_file}")
    print()
    
    # List what's available
    print("Available intelligence:")
    for key in sorted(intelligence_data.keys()):
        size = len(intelligence_data[key])
        print(f"  • {key}: {size:,} chars")

if __name__ == '__main__':
    main()
