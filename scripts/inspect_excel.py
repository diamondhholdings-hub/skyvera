#!/usr/bin/env python3
"""
Inspect Excel file structure to understand layout
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

EXCEL_FILE = "2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx"

project_root = Path(__file__).parent.parent
file_path = project_root / EXCEL_FILE

wb = load_workbook(file_path, data_only=True)

# Check P&Ls - Cloudsense sheet
ws = wb["P&Ls - Cloudsense"]

print("=== First 30 rows, first 15 columns of P&Ls - Cloudsense ===\n")

for row_idx in range(1, 31):
    row_data = []
    for col_idx in range(1, 16):
        val = ws.cell(row_idx, col_idx).value
        if val is not None:
            val_str = str(val)[:30]  # Truncate long values
            row_data.append(f"[{col_idx}]:{val_str}")

    if row_data:
        print(f"Row {row_idx:2d}: {' | '.join(row_data)}")
