#!/usr/bin/env python3
"""
Comprehensive customer analysis from RR Input and NRR Input tabs
"""
from openpyxl import load_workbook
from collections import defaultdict
import json

file_path = "2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx"

print("Loading workbook...")
wb = load_workbook(file_path, data_only=True)

# Function to parse customer data from a sheet
def parse_customer_sheet(sheet_name, revenue_type):
    ws = wb[sheet_name]

    # Find header row (look for "Customer" column)
    header_row_idx = None
    headers = []

    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[1] == 'Customer':  # Customer is in column B (index 1)
            header_row_idx = idx
            headers = list(row)
            break

    if not header_row_idx:
        print(f"Warning: Could not find header row in {sheet_name}")
        return []

    print(f"\n{revenue_type} Headers found at row {header_row_idx}: {headers[:15]}")

    # Parse data rows
    customers = []
    for idx, row in enumerate(ws.iter_rows(values_only=True, min_row=header_row_idx + 1), header_row_idx + 1):
        customer_name = row[1] if len(row) > 1 else None

        # Skip empty rows or summary rows
        if not customer_name or customer_name in ['Customer', None, '']:
            continue

        # Extract key fields
        try:
            customer_data = {
                'customer': customer_name,
                'class': row[2] if len(row) > 2 else None,
                'dept': row[3] if len(row) > 3 else None,
                'account_nature': row[4] if len(row) > 4 else None,
                'ttm_total': float(row[5]) if len(row) > 5 and row[5] is not None and row[5] != '' else 0.0,
                'ttm_avg': float(row[6]) if len(row) > 6 and row[6] is not None and row[6] != '' else 0.0,
                't3m': float(row[7]) if len(row) > 7 and row[7] is not None and row[7] != '' else 0.0,
                'q1_26': float(row[8]) if len(row) > 8 and row[8] is not None and str(row[8]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'q2_26': float(row[9]) if len(row) > 9 and row[9] is not None and str(row[9]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'q3_26': float(row[10]) if len(row) > 10 and row[10] is not None and str(row[10]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'q4_26': float(row[11]) if len(row) > 11 and row[11] is not None and str(row[11]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'fy26_total': float(row[12]) if len(row) > 12 and row[12] is not None and str(row[12]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'notes': row[13] if len(row) > 13 else None,
                'revenue_type': revenue_type,
                'row_num': idx
            }

            # Only add if there's meaningful data
            if customer_data['ttm_total'] > 0 or customer_data['fy26_total'] > 0:
                customers.append(customer_data)
        except (ValueError, TypeError) as e:
            # Skip rows with data issues
            continue

    return customers

# Parse both sheets
print("\n" + "=" * 80)
print("PARSING RR INPUT (Recurring Revenue)")
print("=" * 80)
rr_customers = parse_customer_sheet('RR Input', 'Recurring')

print("\n" + "=" * 80)
print("PARSING NRR INPUT (Non-Recurring Revenue)")
print("=" * 80)
nrr_customers = parse_customer_sheet('NRR Input', 'Non-Recurring')

all_customers = rr_customers + nrr_customers

print(f"\n\nTotal customers parsed:")
print(f"  RR customers: {len(rr_customers)}")
print(f"  NRR customers: {len(nrr_customers)}")
print(f"  Total: {len(all_customers)}")

# ANALYSIS 1: Top customers by revenue
print("\n" + "=" * 80)
print("TOP 15 CUSTOMERS BY FY'26 FORECAST")
print("=" * 80)

sorted_by_fy26 = sorted(all_customers, key=lambda x: x['fy26_total'], reverse=True)
for idx, customer in enumerate(sorted_by_fy26[:15], 1):
    print(f"{idx:2d}. {customer['customer'][:50]:50s} | ${customer['fy26_total']:>12,.0f} | {customer['revenue_type']:13s} | {customer['class'] or 'N/A'}")

# ANALYSIS 2: Revenue concentration
print("\n" + "=" * 80)
print("REVENUE CONCENTRATION ANALYSIS")
print("=" * 80)

total_fy26 = sum(c['fy26_total'] for c in all_customers)
top_10_revenue = sum(c['fy26_total'] for c in sorted_by_fy26[:10])
top_20_revenue = sum(c['fy26_total'] for c in sorted_by_fy26[:20])

print(f"Total FY'26 Forecast: ${total_fy26:,.0f}")
print(f"Top 10 customers: ${top_10_revenue:,.0f} ({top_10_revenue/total_fy26*100:.1f}%)")
print(f"Top 20 customers: ${top_20_revenue:,.0f} ({top_20_revenue/total_fy26*100:.1f}%)")

# ANALYSIS 3: Business Unit breakdown
print("\n" + "=" * 80)
print("REVENUE BY BUSINESS UNIT (FY'26)")
print("=" * 80)

bu_revenue = defaultdict(float)
bu_customer_count = defaultdict(int)

for customer in all_customers:
    class_name = customer['class'] or 'Unknown'
    # Extract BU from class name
    if 'Cloudsense' in class_name:
        bu = 'Cloudsense'
    elif 'Kandy' in class_name:
        bu = 'Kandy'
    elif 'STL' in class_name or 'Software Technology' in class_name:
        bu = 'STL'
    else:
        bu = 'Other'

    bu_revenue[bu] += customer['fy26_total']
    bu_customer_count[bu] += 1

for bu in sorted(bu_revenue.keys(), key=lambda x: bu_revenue[x], reverse=True):
    print(f"{bu:15s}: ${bu_revenue[bu]:>12,.0f} ({bu_revenue[bu]/total_fy26*100:5.1f}%) | {bu_customer_count[bu]:3d} customers")

# ANALYSIS 4: RR vs NRR breakdown
print("\n" + "=" * 80)
print("RECURRING VS NON-RECURRING REVENUE (FY'26)")
print("=" * 80)

rr_total = sum(c['fy26_total'] for c in rr_customers)
nrr_total = sum(c['fy26_total'] for c in nrr_customers)

print(f"Recurring Revenue (RR):     ${rr_total:>12,.0f} ({rr_total/total_fy26*100:5.1f}%)")
print(f"Non-Recurring Revenue (NRR): ${nrr_total:>12,.0f} ({nrr_total/total_fy26*100:5.1f}%)")
print(f"Total:                       ${total_fy26:>12,.0f}")

# ANALYSIS 5: Growth/Decline trends (comparing TTM to FY26)
print("\n" + "=" * 80)
print("GROWTH/DECLINE ANALYSIS (TTM vs FY'26 Forecast)")
print("=" * 80)

growing_customers = []
declining_customers = []

for customer in all_customers:
    if customer['ttm_total'] > 0:  # Only compare customers with historical revenue
        change = customer['fy26_total'] - customer['ttm_total']
        pct_change = (change / customer['ttm_total']) * 100 if customer['ttm_total'] > 0 else 0

        if change > 0:
            growing_customers.append((customer, change, pct_change))
        elif change < 0:
            declining_customers.append((customer, change, pct_change))

growing_customers.sort(key=lambda x: x[1], reverse=True)
declining_customers.sort(key=lambda x: x[1])

print(f"\nTop 10 Growing Customers:")
for idx, (customer, change, pct) in enumerate(growing_customers[:10], 1):
    print(f"{idx:2d}. {customer['customer'][:40]:40s} | +${change:>10,.0f} ({pct:>6.1f}%) | {customer['revenue_type']}")

print(f"\nTop 10 Declining Customers:")
for idx, (customer, change, pct) in enumerate(declining_customers[:10], 1):
    print(f"{idx:2d}. {customer['customer'][:40]:40s} | ${change:>10,.0f} ({pct:>6.1f}%) | {customer['revenue_type']}")

# ANALYSIS 6: Customers at risk (declining revenue or zero FY26)
print("\n" + "=" * 80)
print("CUSTOMERS AT RISK (High TTM, Low/Zero FY'26)")
print("=" * 80)

at_risk = []
for customer in all_customers:
    if customer['ttm_total'] > 50000 and customer['fy26_total'] < customer['ttm_total'] * 0.5:
        at_risk.append(customer)

at_risk.sort(key=lambda x: x['ttm_total'], reverse=True)

print(f"\nFound {len(at_risk)} customers at risk (had >$50K TTM, now <50% of that):")
for idx, customer in enumerate(at_risk[:15], 1):
    print(f"{idx:2d}. {customer['customer'][:40]:40s} | TTM: ${customer['ttm_total']:>10,.0f} | FY'26: ${customer['fy26_total']:>10,.0f}")

# Save data to JSON for dashboard
print("\n" + "=" * 80)
print("SAVING DATA FOR DASHBOARD")
print("=" * 80)

dashboard_data = {
    'summary': {
        'total_fy26': total_fy26,
        'rr_total': rr_total,
        'nrr_total': rr_total,
        'total_customers': len(all_customers),
        'rr_customers': len(rr_customers),
        'nrr_customers': len(nrr_customers),
        'top_10_concentration': (top_10_revenue/total_fy26*100) if total_fy26 > 0 else 0,
        'top_20_concentration': (top_20_revenue/total_fy26*100) if total_fy26 > 0 else 0,
    },
    'top_customers': [
        {
            'customer': c['customer'],
            'fy26_total': c['fy26_total'],
            'revenue_type': c['revenue_type'],
            'class': c['class'],
            'q1': c['q1_26'],
            'q2': c['q2_26'],
            'q3': c['q3_26'],
            'q4': c['q4_26'],
        }
        for c in sorted_by_fy26[:20]
    ],
    'bu_breakdown': [
        {'bu': bu, 'revenue': bu_revenue[bu], 'customers': bu_customer_count[bu]}
        for bu in sorted(bu_revenue.keys(), key=lambda x: bu_revenue[x], reverse=True)
    ],
    'growing_customers': [
        {
            'customer': c[0]['customer'],
            'ttm': c[0]['ttm_total'],
            'fy26': c[0]['fy26_total'],
            'change': c[1],
            'pct_change': c[2]
        }
        for c in growing_customers[:10]
    ],
    'declining_customers': [
        {
            'customer': c[0]['customer'],
            'ttm': c[0]['ttm_total'],
            'fy26': c[0]['fy26_total'],
            'change': c[1],
            'pct_change': c[2]
        }
        for c in declining_customers[:10]
    ],
    'at_risk_customers': [
        {
            'customer': c['customer'],
            'ttm': c['ttm_total'],
            'fy26': c['fy26_total'],
            'class': c['class']
        }
        for c in at_risk[:15]
    ]
}

with open('customer_analysis_data.json', 'w') as f:
    json.dump(dashboard_data, f, indent=2)

print("Data saved to customer_analysis_data.json")
print("\nAnalysis complete!")
