#!/usr/bin/env python3
"""
Check RR Input structure
"""
from openpyxl import load_workbook

file_path = "2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx"
wb = load_workbook(file_path, data_only=True)

ws = wb['RR Input']

print("First 30 rows of RR Input sheet:\n")
for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if idx <= 30:
        print(f"Row {idx}: {row[:15]}")  # First 15 columns
    else:
        break

print(f"\nTotal rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")
