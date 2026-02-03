#!/usr/bin/env python3
"""
Analyze customer data from RR input and NRR input tabs
"""
from openpyxl import load_workbook
import json

file_path = "2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx"

print("Loading workbook...")
wb = load_workbook(file_path, data_only=True)

print(f"\nAvailable sheets: {wb.sheetnames}\n")

# Find the RR and NRR input sheets
rr_sheet_name = None
nrr_sheet_name = None

for sheet_name in wb.sheetnames:
    if 'RR' in sheet_name.upper() and 'INPUT' in sheet_name.upper():
        rr_sheet_name = sheet_name
    if 'NRR' in sheet_name.upper() and 'INPUT' in sheet_name.upper():
        nrr_sheet_name = sheet_name

print(f"RR Input sheet: {rr_sheet_name}")
print(f"NRR Input sheet: {nrr_sheet_name}\n")

# Analyze RR Input sheet
if rr_sheet_name:
    print(f"=" * 80)
    print(f"ANALYZING: {rr_sheet_name}")
    print(f"=" * 80)
    ws = wb[rr_sheet_name]

    print(f"\nFirst 20 rows of data:")
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if idx <= 20:
            print(f"Row {idx}: {row}")
        else:
            break

    print(f"\nTotal rows in sheet: {ws.max_row}")
    print(f"Total columns in sheet: {ws.max_column}")

# Analyze NRR Input sheet
if nrr_sheet_name:
    print(f"\n{'=' * 80}")
    print(f"ANALYZING: {nrr_sheet_name}")
    print(f"=" * 80)
    ws = wb[nrr_sheet_name]

    print(f"\nFirst 20 rows of data:")
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if idx <= 20:
            print(f"Row {idx}: {row}")
        else:
            break

    print(f"\nTotal rows in sheet: {ws.max_row}")
    print(f"Total columns in sheet: {ws.max_column}")

print("\nAnalysis complete!")
