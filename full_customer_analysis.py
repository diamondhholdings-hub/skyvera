#!/usr/bin/env python3
"""
Complete customer analysis from both RR Input and NRR Input tabs
"""
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import json

file_path = "2025-12-11 Skyvera - Budget - Q1'26 - For Todd.xlsx"

print("Loading workbook...")
wb = load_workbook(file_path, data_only=True)

# Parse RR Input (Recurring Revenue - subscription based)
def parse_rr_input():
    ws = wb['RR Input']

    print("\nParsing RR Input...")
    customers = []

    # Header is at row 10
    for idx, row in enumerate(ws.iter_rows(values_only=True, min_row=11), 11):
        company = row[0]
        customer_name = row[1]

        # Skip empty rows
        if not customer_name:
            continue

        try:
            arr_current = float(row[6]) if row[6] is not None and str(row[6]).replace('.', '').replace('-', '').isdigit() else 0.0
            will_renew = str(row[9]) if row[9] else 'Unknown'
            upsell_pct = float(row[10]) if row[10] is not None and str(row[10]).replace('.', '').replace('-', '').isdigit() else 0.0
            projected_arr = float(row[11]) if row[11] is not None and str(row[11]).replace('.', '').replace('-', '').isdigit() else 0.0

            customer_data = {
                'company': company,
                'customer': customer_name,
                'end_user': row[2],
                'subscription_id': row[3],
                'arr_current': arr_current,
                'renewal_qtr': row[8],
                'will_renew': will_renew,
                'upsell_pct': upsell_pct,
                'projected_arr': projected_arr,
                'revenue_type': 'Recurring',
                'row_num': idx
            }

            # Only add if there's meaningful ARR
            if customer_data['arr_current'] > 0 or customer_data['projected_arr'] > 0:
                customers.append(customer_data)

        except (ValueError, TypeError, IndexError) as e:
            continue

    return customers

# Parse NRR Input (Non-Recurring Revenue)
def parse_nrr_input():
    ws = wb['NRR Input']

    customers = []
    # Header is at row 5
    for idx, row in enumerate(ws.iter_rows(values_only=True, min_row=6), 6):
        customer_name = row[1]

        # Skip empty rows or summary rows
        if not customer_name or customer_name in ['Customer', None, '']:
            continue

        try:
            customer_data = {
                'customer': customer_name,
                'class': row[2] if len(row) > 2 else None,
                'dept': row[3] if len(row) > 3 else None,
                'ttm_total': float(row[5]) if len(row) > 5 and row[5] is not None and row[5] != '' else 0.0,
                'ttm_avg': float(row[6]) if len(row) > 6 and row[6] is not None and row[6] != '' else 0.0,
                't3m': float(row[7]) if len(row) > 7 and row[7] is not None and row[7] != '' else 0.0,
                'q1_26': float(row[8]) if len(row) > 8 and row[8] is not None and str(row[8]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'q2_26': float(row[9]) if len(row) > 9 and row[9] is not None and str(row[9]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'q3_26': float(row[10]) if len(row) > 10 and row[10] is not None and str(row[10]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'q4_26': float(row[11]) if len(row) > 11 and row[11] is not None and str(row[11]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'fy26_total': float(row[12]) if len(row) > 12 and row[12] is not None and str(row[12]).replace('.', '').replace('-', '').isdigit() else 0.0,
                'notes': row[13] if len(row) > 13 else None,
                'revenue_type': 'Non-Recurring',
                'row_num': idx
            }

            # Only add if there's meaningful data
            if customer_data['ttm_total'] > 0 or customer_data['fy26_total'] > 0:
                customers.append(customer_data)
        except (ValueError, TypeError) as e:
            continue

    return customers

print("=" * 80)
print("PARSING CUSTOMER DATA")
print("=" * 80)

rr_customers = parse_rr_input()
nrr_customers = parse_nrr_input()

print(f"\nParsed:")
print(f"  RR subscriptions: {len(rr_customers)}")
print(f"  NRR customers: {len(nrr_customers)}")

# RR ANALYSIS
print("\n" + "=" * 80)
print("RECURRING REVENUE (RR) ANALYSIS")
print("=" * 80)

total_arr_current = sum(c['arr_current'] for c in rr_customers)
total_arr_projected = sum(c['projected_arr'] for c in rr_customers)
arr_growth = total_arr_projected - total_arr_current

print(f"\nARR Summary:")
print(f"  Current ARR (as of 11/30/2025): ${total_arr_current:,.0f}")
print(f"  Projected ARR:                   ${total_arr_projected:,.0f}")
print(f"  Growth/Decline:                  ${arr_growth:,.0f} ({arr_growth/total_arr_current*100:.1f}%)")
print(f"  Quarterly RR (ARR/4):            ${total_arr_projected/4:,.0f}")

# Top customers by ARR
print("\n" + "=" * 80)
print("TOP 15 CUSTOMERS BY CURRENT ARR")
print("=" * 80)

sorted_rr = sorted(rr_customers, key=lambda x: x['arr_current'], reverse=True)
for idx, customer in enumerate(sorted_rr[:15], 1):
    renewal_status = "✓" if 'Yes' in str(customer['will_renew']) else "✗"
    print(f"{idx:2d}. {customer['customer'][:40]:40s} | ${customer['arr_current']:>11,.0f} | {customer['company']:10s} | {renewal_status}")

# Renewal risk analysis
print("\n" + "=" * 80)
print("RENEWAL RISK ANALYSIS")
print("=" * 80)

at_risk_subscriptions = [c for c in rr_customers if 'No' in str(c['will_renew']) or 'SF' in str(c['will_renew'])]
at_risk_arr = sum(c['arr_current'] for c in at_risk_subscriptions)

print(f"\nSubscriptions Not Renewing: {len(at_risk_subscriptions)}")
print(f"ARR at Risk: ${at_risk_arr:,.0f} ({at_risk_arr/total_arr_current*100:.1f}% of total ARR)")

print(f"\nTop 10 At-Risk Subscriptions:")
at_risk_sorted = sorted(at_risk_subscriptions, key=lambda x: x['arr_current'], reverse=True)
for idx, customer in enumerate(at_risk_sorted[:10], 1):
    print(f"{idx:2d}. {customer['customer'][:40]:40s} | ${customer['arr_current']:>11,.0f} | {customer['renewal_qtr']}")

# Business unit breakdown (RR)
print("\n" + "=" * 80)
print("RR BY BUSINESS UNIT")
print("=" * 80)

bu_arr = defaultdict(float)
bu_count = defaultdict(int)

for customer in rr_customers:
    company = customer['company'] or 'Unknown'
    if 'Cloudsense' in company:
        bu = 'Cloudsense'
    elif 'Kandy' in company:
        bu = 'Kandy'
    elif 'STL' in company:
        bu = 'STL'
    elif 'NewNet' in company:
        bu = 'NewNet'
    else:
        bu = 'Other'

    bu_arr[bu] += customer['projected_arr']
    bu_count[bu] += 1

for bu in sorted(bu_arr.keys(), key=lambda x: bu_arr[x], reverse=True):
    print(f"{bu:15s}: ${bu_arr[bu]:>12,.0f} ({bu_arr[bu]/total_arr_projected*100:5.1f}%) | {bu_count[bu]:3d} subscriptions")

# NRR ANALYSIS
print("\n" + "=" * 80)
print("NON-RECURRING REVENUE (NRR) ANALYSIS")
print("=" * 80)

total_nrr_fy26 = sum(c['fy26_total'] for c in nrr_customers)
total_nrr_ttm = sum(c['ttm_total'] for c in nrr_customers)

print(f"\nNRR Summary:")
print(f"  TTM Total:  ${total_nrr_ttm:,.0f}")
print(f"  FY'26 Plan: ${total_nrr_fy26:,.0f}")
print(f"  Change:     ${total_nrr_fy26 - total_nrr_ttm:,.0f} ({(total_nrr_fy26 - total_nrr_ttm)/total_nrr_ttm*100:.1f}%)")

print("\n" + "=" * 80)
print("TOP 10 NRR CUSTOMERS (FY'26 Forecast)")
print("=" * 80)

sorted_nrr = sorted(nrr_customers, key=lambda x: x['fy26_total'], reverse=True)
for idx, customer in enumerate(sorted_nrr[:10], 1):
    class_name = (customer['class'] or 'Unknown')[:20]
    print(f"{idx:2d}. {customer['customer'][:40]:40s} | ${customer['fy26_total']:>11,.0f} | {class_name}")

# COMBINED ANALYSIS
print("\n" + "=" * 80)
print("COMBINED REVENUE SUMMARY (FY'26)")
print("=" * 80)

total_rr_quarterly = total_arr_projected / 4
total_revenue_fy26 = (total_rr_quarterly * 4) + total_nrr_fy26

print(f"\nRecurring Revenue (RR):     ${total_rr_quarterly * 4:>12,.0f} ({(total_rr_quarterly * 4)/total_revenue_fy26*100:5.1f}%)")
print(f"Non-Recurring Revenue (NRR): ${total_nrr_fy26:>12,.0f} ({total_nrr_fy26/total_revenue_fy26*100:5.1f}%)")
print(f"Total Revenue FY'26:         ${total_revenue_fy26:>12,.0f}")

# Customer concentration (combined)
print("\n" + "=" * 80)
print("CUSTOMER CONCENTRATION ANALYSIS")
print("=" * 80)

# Create combined customer list for concentration analysis
combined_customers = []

# Add RR customers (annualized)
for c in rr_customers:
    combined_customers.append({
        'customer': c['customer'],
        'annual_revenue': c['projected_arr'],
        'type': 'Recurring',
        'company': c.get('company', 'Unknown')
    })

# Add NRR customers
for c in nrr_customers:
    combined_customers.append({
        'customer': c['customer'],
        'annual_revenue': c['fy26_total'],
        'type': 'Non-Recurring',
        'company': c.get('class', 'Unknown')
    })

# Aggregate by customer name (some may appear in both RR and NRR)
customer_totals = defaultdict(lambda: {'revenue': 0, 'types': set(), 'company': ''})
for c in combined_customers:
    customer_totals[c['customer']]['revenue'] += c['annual_revenue']
    customer_totals[c['customer']]['types'].add(c['type'])
    customer_totals[c['customer']]['company'] = c['company']

aggregated_customers = [
    {'customer': name, 'revenue': data['revenue'], 'types': ','.join(data['types']), 'company': data['company']}
    for name, data in customer_totals.items()
]

sorted_combined = sorted(aggregated_customers, key=lambda x: x['revenue'], reverse=True)

print("\nTop 20 Customers (All Revenue Types):")
for idx, c in enumerate(sorted_combined[:20], 1):
    print(f"{idx:2d}. {c['customer'][:40]:40s} | ${c['revenue']:>11,.0f} | {c['types']}")

top_10_rev = sum(c['revenue'] for c in sorted_combined[:10])
top_20_rev = sum(c['revenue'] for c in sorted_combined[:20])
total_rev = sum(c['revenue'] for c in aggregated_customers)

print(f"\nConcentration Metrics:")
print(f"  Top 10 customers: ${top_10_rev:,.0f} ({top_10_rev/total_rev*100:.1f}%)")
print(f"  Top 20 customers: ${top_20_rev:,.0f} ({top_20_rev/total_rev*100:.1f}%)")

# Save data for dashboard
print("\n" + "=" * 80)
print("SAVING DATA FOR DASHBOARD")
print("=" * 80)

dashboard_data = {
    'summary': {
        'total_revenue_fy26': total_revenue_fy26,
        'rr_annual': total_arr_projected,
        'nrr_annual': total_nrr_fy26,
        'arr_current': total_arr_current,
        'arr_growth': arr_growth,
        'arr_at_risk': at_risk_arr,
        'total_rr_subscriptions': len(rr_customers),
        'total_nrr_customers': len(nrr_customers),
        'at_risk_subscriptions': len(at_risk_subscriptions),
        'top_10_concentration_pct': (top_10_rev/total_rev*100) if total_rev > 0 else 0,
        'top_20_concentration_pct': (top_20_rev/total_rev*100) if total_rev > 0 else 0,
    },
    'top_rr_customers': [
        {
            'customer': c['customer'],
            'arr_current': c['arr_current'],
            'projected_arr': c['projected_arr'],
            'company': c['company'],
            'will_renew': c['will_renew'],
            'renewal_qtr': c['renewal_qtr']
        }
        for c in sorted_rr[:15]
    ],
    'at_risk_rr': [
        {
            'customer': c['customer'],
            'arr_current': c['arr_current'],
            'renewal_qtr': c['renewal_qtr'],
            'company': c['company']
        }
        for c in at_risk_sorted[:10]
    ],
    'top_nrr_customers': [
        {
            'customer': c['customer'],
            'fy26_total': c['fy26_total'],
            'ttm_total': c['ttm_total'],
            'class': c['class'],
            'q1': c['q1_26'],
            'q2': c['q2_26'],
            'q3': c['q3_26'],
            'q4': c['q4_26'],
        }
        for c in sorted_nrr[:15]
    ],
    'bu_breakdown_rr': [
        {'bu': bu, 'arr': bu_arr[bu], 'subscriptions': bu_count[bu]}
        for bu in sorted(bu_arr.keys(), key=lambda x: bu_arr[x], reverse=True)
    ],
    'top_combined_customers': [
        {
            'customer': c['customer'],
            'revenue': c['revenue'],
            'types': c['types'],
            'company': c['company']
        }
        for c in sorted_combined[:20]
    ],
    'nrr_declining': [
        {
            'customer': c['customer'],
            'ttm': c['ttm_total'],
            'fy26': c['fy26_total'],
            'change': c['fy26_total'] - c['ttm_total']
        }
        for c in sorted(nrr_customers, key=lambda x: (x['fy26_total'] - x['ttm_total']))[:10]
    ]
}

with open('customer_dashboard_data.json', 'w') as f:
    json.dump(dashboard_data, f, indent=2)

print("Data saved to customer_dashboard_data.json")
print("\nAnalysis complete!")
